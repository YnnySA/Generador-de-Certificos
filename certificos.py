import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import sqlite3
import os
from datetime import datetime

# Inicializar session state 
if 'facturas_rows' not in st.session_state:
    st.session_state.facturas_rows = 1
    
# --- NUEVO: Inicializar estado para los filtros de búsqueda avanzada ---
if 'filtros_aplicados' not in st.session_state:
    st.session_state.filtros_aplicados = False
if 'filtro_obras' not in st.session_state:
    st.session_state.filtro_obras = []
if 'filtro_estado' not in st.session_state:
    st.session_state.filtro_estado = []
if 'filtro_fecha_inicio' not in st.session_state:
    st.session_state.filtro_fecha_inicio = None
if 'filtro_fecha_fin' not in st.session_state:
    st.session_state.filtro_fecha_fin = None
if 'filtro_contratista' not in st.session_state:
    st.session_state.filtro_contratista = ""

# --- NUEVA FUNCIÓN DE CALLBACK PARA NAVEGACIÓN ---
def go_to_page(page_name):
    """Función para navegar a una página específica usando query params."""
    st.query_params.page = page_name
    st.rerun()
# --- FIN NUEVO ---

# Configuración de la base de datos y directorios
DB_NAME = "certificados.db"
EXCEL_TEMPLATES_DIR = "data"
CERTIFICADOS_DIR = "certificados_generados"

# Crear directorios necesarios
os.makedirs(EXCEL_TEMPLATES_DIR, exist_ok=True)
os.makedirs(CERTIFICADOS_DIR, exist_ok=True)

# Inicializar la base de datos
def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    
    # Crear tabla para obras
    c.execute('''CREATE TABLE IF NOT EXISTS obras (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT UNIQUE NOT NULL,
        codigo INTEGER NOT NULL,
        aprobacion TEXT NOT NULL
    )''')
    
    # Crear tabla para certificados (con UNIQUE constraint correcta y nuevos campos para estado)
    c.execute('''CREATE TABLE IF NOT EXISTS certificados (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        numero_certificado INTEGER NOT NULL,
        obra_id INTEGER,
        fecha DATE NOT NULL,
        contrato TEXT,
        contratista TEXT,
        valor_contrato REAL,
        valor_pagado REAL,
        total_facturas REAL,
        archivo_path TEXT,
        fecha_generacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        -- Nuevos campos para el estado
        estado TEXT DEFAULT 'Activo', -- 'Activo', 'Revertido', 'Cancelado'
        comentario_estado TEXT,
        FOREIGN KEY (obra_id) REFERENCES obras (id),
        UNIQUE(obra_id, numero_certificado)
    )''')
    
    # Crear tabla para facturas
    c.execute('''CREATE TABLE IF NOT EXISTS facturas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        certificado_id INTEGER,
        proveedor TEXT NOT NULL,
        numero_factura TEXT NOT NULL,
        importe REAL NOT NULL,
        codigo TEXT,
        FOREIGN KEY (certificado_id) REFERENCES certificados (id)
    )''')
    
    # Insertar obras iniciales si no existen
    obras_iniciales = [
        ('Mejoras Cayo Saetía', 759, 'A 37-018-15'),
        ('Marina Cayo Saetía', 677, 'A 37-024-19'),
        ('Viviendas Mayarí', 699, 'A 37-037-20'),
        ('Delfinario Cayo Saetía', 605, 'A 37-025-19'),
        ('Canal Dumois', 872, 'A 37-038-21')
    ]
    
    for nombre, codigo, aprobacion in obras_iniciales:
        try:
            c.execute("INSERT INTO obras (nombre, codigo, aprobacion) VALUES (?, ?, ?)",
                     (nombre, codigo, aprobacion))
        except sqlite3.IntegrityError:
            pass  # La obra ya existe
    
    conn.commit()
    conn.close()

# Función para obtener el siguiente número de certificado PARA UNA OBRA ESPECÍFICA
def get_next_certificado_number_por_obra(obra_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    # Obtiene el máximo número de certificado para la obra dada
    c.execute("SELECT MAX(numero_certificado) FROM certificados WHERE obra_id = ?", (obra_id,))
    result = c.fetchone()[0]
    conn.close()
    # Si no hay certificados para esta obra, el siguiente es 1
    return (result or 0) + 1

# Función para obtener todas las obras
def get_all_obras():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT id, nombre, codigo, aprobacion FROM obras ORDER BY nombre")
    obras = c.fetchall()
    conn.close()
    return obras

# Función para obtener un certificado por ID (incluyendo estado)
def get_certificado_by_id(certificado_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""SELECT c.*, o.nombre as obra_nombre, o.codigo as obra_codigo, o.aprobacion 
                 FROM certificados c 
                 JOIN obras o ON c.obra_id = o.id 
                 WHERE c.id = ?""", (certificado_id,))
    certificado = c.fetchone()
    conn.close()
    return certificado

# Función para obtener facturas de un certificado
def get_facturas_by_certificado_id(certificado_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT proveedor, numero_factura, importe, codigo FROM facturas WHERE certificado_id = ? ORDER BY id", 
              (certificado_id,))
    facturas = c.fetchall()
    conn.close()
    return facturas

# Función para actualizar un certificado (incluyendo estado)
def update_certificado(certificado_id, fecha, contrato, contratista, valor_contrato, valor_pagado, total_facturas, estado, comentario_estado):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""UPDATE certificados 
                 SET fecha = ?, contrato = ?, contratista = ?, valor_contrato = ?, valor_pagado = ?, total_facturas = ?,
                     estado = ?, comentario_estado = ?
                 WHERE id = ?""",
              (fecha, contrato, contratista, valor_contrato, valor_pagado, total_facturas, estado, comentario_estado, certificado_id))
    conn.commit()
    conn.close()

# Función para actualizar facturas de un certificado
def update_facturas(certificado_id, facturas_data):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    
    # Eliminar facturas existentes
    c.execute("DELETE FROM facturas WHERE certificado_id = ?", (certificado_id,))
    
    # Insertar nuevas facturas
    for factura in facturas_data:
        c.execute("""INSERT INTO facturas (certificado_id, proveedor, numero_factura, importe, codigo) 
                     VALUES (?, ?, ?, ?, ?)""",
                  (certificado_id, factura['proveedor'], factura['factura'], factura['importe'], factura['codigo']))
    
    conn.commit()
    conn.close()

# Función para eliminar un certificado
def delete_certificado(certificado_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    
    # Eliminar primero las facturas asociadas
    c.execute("DELETE FROM facturas WHERE certificado_id = ?", (certificado_id,))
    
    # Luego eliminar el certificado
    c.execute("DELETE FROM certificados WHERE id = ?", (certificado_id,))
    
    conn.commit()
    conn.close()

# Función para obtener certificados por obra (incluyendo estado)
def get_certificados_by_obra(obra_id=None):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    
    if obra_id:
        c.execute("""SELECT c.*, o.nombre as obra_nombre, o.codigo as obra_codigo
                     FROM certificados c 
                     JOIN obras o ON c.obra_id = o.id 
                     WHERE c.obra_id = ? 
                     ORDER BY c.numero_certificado DESC""", (obra_id,))
    else:
        c.execute("""SELECT c.*, o.nombre as obra_nombre, o.codigo as obra_codigo
                     FROM certificados c 
                     JOIN obras o ON c.obra_id = o.id 
                     ORDER BY o.nombre, c.numero_certificado DESC""")
    
    certificados = c.fetchall()
    conn.close()
    return certificados

# --- FUNCIÓN DE BÚSQUEDA AVANZADA ---
def buscar_certificados_con_filtros(obras_ids=None, estados=None, fecha_inicio=None, fecha_fin=None, contratista_texto=None):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    
    # Consulta base con JOIN para obtener el nombre de la obra
    query = """
        SELECT c.*, o.nombre as obra_nombre, o.codigo as obra_codigo
        FROM certificados c 
        JOIN obras o ON c.obra_id = o.id 
        WHERE 1=1
    """
    params = []
    
    # Añadir filtros dinámicamente si se proporcionan
    if obras_ids:
        placeholders = ','.join(['?'] * len(obras_ids))
        query += f" AND c.obra_id IN ({placeholders})"
        params.extend(obras_ids)
        
    if estados:
        placeholders = ','.join(['?'] * len(estados))
        query += f" AND c.estado IN ({placeholders})"
        params.extend(estados)
        
    if fecha_inicio:
        query += " AND c.fecha >= ?"
        params.append(fecha_inicio)
        
    if fecha_fin:
        query += " AND c.fecha <= ?"
        params.append(fecha_fin)
        
    if contratista_texto:
        query += " AND c.contratista LIKE ?"
        params.append(f'%{contratista_texto}%')
        
    query += " ORDER BY o.nombre, c.numero_certificado DESC"
    
    c.execute(query, params)
    certificados = c.fetchall()
    conn.close()
    return certificados

# Función para agregar una nueva fila
def agregar_fila():
    st.session_state.facturas_rows += 1

# Función para eliminar la última fila
def eliminar_fila():
    if st.session_state.facturas_rows > 1:
        st.session_state.facturas_rows -= 1

# Función para validar campos obligatorios
def validar_campos_obligatorios(fecha, obra_seleccionada, facturas_data):
    errores = []
    
    # Validar fecha
    if not fecha:
        errores.append("❌ Debe seleccionar una fecha")
    
    # Validar que se haya seleccionado una obra
    if not obra_seleccionada:
        errores.append("❌ Debe seleccionar una obra")
    
    # Validar facturas
    if len(facturas_data) == 0:
        errores.append("❌ Debe agregar al menos una factura")
    else:
        for i, factura in enumerate(facturas_data):
            factura_num = i + 1
            if not factura['proveedor'].strip():
                errores.append(f"❌ Factura No {factura_num}: Debe ingresar el proveedor")
            if not factura['factura'].strip():
                errores.append(f"❌ Factura No {factura_num}: Debe ingresar el número de factura")
            if factura['importe'] <= 0:
                errores.append(f"❌ Factura No {factura_num}: El importe debe ser mayor que 0")
    
    return errores

# Función para crear el informe en Excel (actualizada para mostrar estado)
def generar_informe_excel(datos, numero_certificado):
    try:
        # Cargar la plantilla 
        wb = load_workbook('data/ejemplo.xlsx')
        ws = wb.active
        
        # Llenar los datos en las celdas correspondientes
        
        # Colocar el número de certificado consecutivo 
        ws['E11'] = numero_certificado
        
        # Fecha 
        if datos['fecha']:
            ws['E6'] = datos['fecha']
            
        # Contrato 
        if datos['contrato']:
            ws['B13'] = datos['contrato']
            
        # Contratista 
        if datos['contratista']:
            ws['B16'] = datos['contratista']
            
        # Obra 
        if datos['obra']:
            ws['B16'] = datos['obra']
        
        if datos['codigo_obra'] is not None: # Verificar que el código de obra exista
            ws['E18'] = f"{datos['codigo_obra']}"
            
        # Aprobación (celda A19 en la imagen)
        if datos['aprobacion']:
            ws['B18'] = datos['aprobacion']
            ws['C32'] = datos['aprobacion']
            
        # Valor total contrato (celda A23 en la imagen)
        if datos['valor_contrato']:
            ws['C20'] = f"{datos['valor_contrato']:,.2f}"
            
        # Valor pagado (celda A25 en la imagen)
        if datos['valor_pagado']:
            ws['C22'] = f"{datos['valor_pagado']:,.2f}"
            
        # Insertar las facturas comenzando desde la fila 28 (ajusta según tu plantilla)
        fila_inicio_facturas = 26
        for i, factura in enumerate(datos['facturas']):
            fila = fila_inicio_facturas + i
            if fila <= ws.max_row:
                ws[f'A{fila}'] = factura['proveedor']
                ws[f'C{fila}'] = factura['factura']
                ws[f'E{fila}'] = factura['importe']
                ws[f'F{fila}'] = factura['codigo']
            else:
                # Si hay más filas, agregarlas
                ws.append([factura['proveedor'], factura['factura'], factura['importe'], factura['codigo']])
        
        # Total de facturas (celda E35 en la imagen, aproximadamente)
        ws['E32'] = f"{datos['total_facturas']:,.2f} CUP"
        
        # --- NUEVO: Mostrar estado y comentario en el informe ---
        if 'estado' in datos and datos['estado'] != 'Activo':
            from openpyxl.styles import Font, Alignment, Border, Side
            
            # Para asegurar que se vea bien, fusionamos celdas
            celda_inicio = 'B40'
            celda_fin = 'F42'
            ws.merge_cells(f'{celda_inicio}:{celda_fin}')
            cell = ws[celda_inicio]
            cell.value = f"⚠️ ESTADO DEL CERTIFICADO: {datos['estado']}\n\n📝 Comentario: {datos.get('comentario_estado', 'Ninguno')}"
            cell.font = Font(bold=True, color="FF0000", size=12)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Agregar un borde rojo para hacerlo más visible
            thin_border = Border(
                left=Side(style='thin', color='FF0000'),
                right=Side(style='thin', color='FF0000'),
                top=Side(style='thin', color='FF0000'),
                bottom=Side(style='thin', color='FF0000')
            )
            for row in ws[f'{celda_inicio}:{celda_fin}']:
                for c in row:
                    c.border = thin_border
                    
        # --- FIN NUEVO ---
        
        # Guardar el archivo en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    except Exception as e:
        st.error(f"Error al generar el informe: {str(e)}")
        return None

# Función para guardar certificado en la base de datos (incluyendo estado por defecto)
def guardar_certificado_db(numero_certificado, obra_id, fecha, contrato, contratista, 
                          valor_contrato, valor_pagado, total_facturas, facturas_data, archivo_path):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    
    try:
        # Insertar certificado con el número específico por obra y estado por defecto 'Activo'
        c.execute("""INSERT INTO certificados 
                     (numero_certificado, obra_id, fecha, contrato, contratista, valor_contrato, 
                      valor_pagado, total_facturas, archivo_path, estado, comentario_estado) 
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                  (numero_certificado, obra_id, fecha, contrato, contratista, 
                   valor_contrato, valor_pagado, total_facturas, archivo_path, 'Activo', None))
        
        certificado_id = c.lastrowid
        
        # Insertar facturas
        for factura in facturas_data:
            c.execute("""INSERT INTO facturas (certificado_id, proveedor, numero_factura, importe, codigo) 
                         VALUES (?, ?, ?, ?, ?)""",
                      (certificado_id, factura['proveedor'], factura['factura'], 
                       factura['importe'], factura['codigo']))
        
        conn.commit()
        conn.close()
        return certificado_id
    except sqlite3.IntegrityError as e:
        conn.rollback()
        conn.close()
        # Relanzar la excepción para que se maneje en el lugar de llamada
        raise e

# Inicializar la base de datos
init_db()

# ==================== INTERFAZ DE USUARIO ====================

# Configuración del estilo de la aplicación
st.set_page_config(
    page_title="Certificados de Obras",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Sidebar para navegación
with st.sidebar:
    st.title("🏗️ _CERTIFICADOS")
    st.markdown("---")
    
    # --- NUEVA LÓGICA DE NAVEGACIÓN CON QUERY PARAMS ---
    
    # Definir las páginas disponibles
    pages = {
        "🏠 Crear Nuevo Certificado": "crear",
        "📋 Ver Certificados": "ver",
        "✏️ Editar Certificado": "editar"
    }
    
    # Obtener la página actual de los query params, por defecto es "crear"
    current_page = st.query_params.get("page", "crear")
    
    # Encontrar el índice de la página actual para el radio button
    try:
        index = list(pages.values()).index(current_page)
    except ValueError:
        index = 0 # Por defecto a la primera página si el query param es inválido
    
    # El radio button ahora es solo para mostrar, no para controlar el estado
    selected_page_name = st.radio(
        "Seleccione una opción:",
        options=list(pages.keys()),
        index=index
    )
    
    # Si el usuario selecciona una opción diferente en el radio, actualizamos el query param
    if pages[selected_page_name] != current_page:
        go_to_page(pages[selected_page_name])
        
    # Determinar la opción del menú basada en el query param
    menu_opcion = selected_page_name

    # --- FIN NUEVA LÓGICA ---
      
    # Información adicional en el sidebar
    st.markdown("---")
    st.info("Sistema de gestión de certificados para obras de construcción")

if menu_opcion == "🏠 Crear Nuevo Certificado":
    st.title("📄 Crear Nuevo Certificado")
    
    # Contenedor para el encabezado
    with st.container():
        encabezado_superior = st.columns([0.20, 0.70, 0.20], vertical_alignment="bottom")

        with encabezado_superior[0]:
            # Verificar si existe el logo antes de mostrarlo
            if os.path.exists('logo.png'):
                st.image('logo.png', width=150)
            else:
                st.markdown("🏢")
        with encabezado_superior[1]:
            st.header('Inmobiliaria ALMEST \n UBI Obras Varias')
        with encabezado_superior[2]:
            fecha = st.date_input('Fecha', format='DD/MM/YYYY', value=datetime.now().date())

    st.write('Certificamos que los valores de las facturas que se relacionan corresponden a los documentos legales debidamente autorizados y que se ajustan a la obra de referencia.')
    
    # Sección de información del contrato
    st.subheader("📋 Información del Contrato")
    with st.container():
        col1, col2 = st.columns(2)
        with col1:
            contrato = st.text_input('No. Contrato')
        with col2:
            contratista = st.text_input('Contratista')

    # Sección de información de la obra
    st.subheader("🏗️ Información de la Obra")
    with st.container():
        # Obtener obras de la base de datos
        data_obras = pd.DataFrame(
            {
            "Obras": ['Mejoras Cayo Saetía', 'Marina Cayo Saetía', 'Viviendas Mayarí', 'Delfinario Cayo Saetía', 'Canal Dumois'], 
            "Código de Obra": [759, 677, 699, 605, 872],
            "Aprobación": ['A 37-018-15', 'A 37-024-19', 'A 37-037-20', 'A 37-025-19', 'A 37-038-21'],   
            })

        obras = st.selectbox('Obra',data_obras["Obras"], index=None, placeholder="Despliegue y seleccione una Obra")

        # Variables para almacenar datos de la obra seleccionada
        codigo_obra = None
        nombre_obra = None
        aprobacion = None
        obra_id = None

        # Cuando se selecciona una obra
        if obras:
            # Filtrar el DataFrame por la obra seleccionada
            obra_seleccionada = data_obras[data_obras["Obras"] == obras]
            
            # Verificar que se encontró la obra
            if not obra_seleccionada.empty:
                # Obtener el código de obra seleccionada
                codigo_obra = obra_seleccionada["Código de Obra"].iloc[0]
                nombre_obra = obra_seleccionada["Obras"].iloc[0]
                
                # Obtener ID de la obra de la base de datos
                obras_db = get_all_obras()
                for obra_db in obras_db:
                    if obra_db[1] == nombre_obra and obra_db[2] == codigo_obra:
                        obra_id = obra_db[0]
                        break
                
                # Mostrar información de la obra seleccionada
                st.info(f"**Obra seleccionada:** {codigo_obra} - {nombre_obra}")
                
                # Obtener la Aprobación de la obra en cuestión
                aprobacion = obra_seleccionada["Aprobación"].iloc[0]  
                st.info(f"**Aprobación:** {aprobacion}")
                st.info(f"**Código de Obra:** {codigo_obra}02")
            else:
                st.warning("No se encontraron datos para esta obra")

    # Sección de valores económicos
    st.subheader("💰 Valores Económicos")
    with st.container():
        col1, col2 = st.columns(2)
        with col1:
            valor_contrato = st.number_input(
                'Valor Total del Contrato (CUP):',
                min_value=0.0,
                max_value=1000000000.0,
                value=0.0,
                step=1000.0,
                format="%.2f",
                help="Introduzca el monto total del contrato en CUP"
            )
        with col2:
            valor_pagado = st.number_input(
                'Valor Total Certificado (CUP):',
                min_value=0.0,
                max_value=1000000000.0,
                value=0.0,
                step=1000.0,
                format="%.2f",
                help="Introduzca el monto total del contrato en CUP"
            )

    # Sección de facturas
    st.subheader("📋 Facturas")
    with st.container():
        # Botones para agregar/eliminar filas
        col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 3])
        with col_btn1:
            st.button("➕ Agregar Factura", on_click=agregar_fila, type="primary")
        with col_btn2:
            st.button("➖ Eliminar Factura", on_click=eliminar_fila, type="secondary")
        with col_btn3:
            st.write("")

        st.divider()

        # Listas para almacenar datos de facturas
        facturas_data = []

        # Crear las filas de facturas
        for i in range(st.session_state.facturas_rows):
            st.markdown(f"**Factura No {i+1}**")
            proovedores, facturas, importe, codigo = st.columns([2, 2, 1.5, 1])
            
            with proovedores:
                proveedor_val = st.text_input(f"Proveedor", key=f"proveedor_{i}")
                
            with facturas:
                factura_val = st.text_input(f"Factura", key=f"factura_{i}")
                
            with importe:
                importe_factura = st.number_input(f"Importe", 
                                                format="%.2f", 
                                                step=1000.0,
                                                key=f"importe_{i}")
                # Mostrar importe formateado
                if importe_factura > 0:
                    st.caption(f"{importe_factura:,.2f} CUP")
            
            with codigo:
                codigo_val = st.text_input(f"Código", key=f"codigo_{i}")
            
            # Almacenar datos de la factura
            facturas_data.append({
                'proveedor': proveedor_val,
                'factura': factura_val,
                'importe': importe_factura,
                'codigo': codigo_val
            })
            
            st.divider()

        # Calcular y mostrar el total de todas las facturas
        total_facturas = sum(f['importe'] for f in facturas_data)

        # Mostrar el total formateado
        st.markdown("---")
        col_total1, col_total2, col_total3 = st.columns([1.1, 2, 2])
        with col_total1:
            st.write("")
        with col_total2:
            st.markdown("**TOTAL DE FACTURAS:**")
        with col_total3:
            st.markdown(f"**{total_facturas:,.2f} CUP**")
            st.divider()

    # Sección de firmas
    st.subheader("✍️ Firmas")
    with st.container():
        st.write(" ")
        pie1, _, pie3 = st.columns([1, 0.5, 1.5])

        with pie1:
            st.write("**Firma:**")
            st.write("Especialista en Inversiones")
            st.write("**Ing. Yenny Sánchez Aguilar**")
            
        with pie3:
            st.write("**Firma:**")
            st.write("Jefe de Grupo Técnico UBI Obras Varias.")
            st.write("**Ing. Osvaldo Sánchez Breff**")

    # Botón para generar el informe
    st.markdown("---")
    with st.container():
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("📄 Generar Informe en Excel", type="primary", use_container_width=True):
                # Validar campos obligatorios
                errores = validar_campos_obligatorios(fecha, obras, facturas_data)
                
                if errores:
                    # Mostrar errores
                    st.error("🚨 Por favor corrija los siguientes errores antes de generar el informe:")
                    for error in errores:
                        st.write(error)
                else:
                    # Verificar que se haya seleccionado una obra para obtener su ID
                    if not obra_id:
                         st.error("❌ Error: No se pudo identificar la obra seleccionada.")
                    else:
                        # Obtener número de certificado consecutivo PARA LA OBRA SELECCIONADA
                        numero_certificado = get_next_certificado_number_por_obra(obra_id)
                        
                        # Recopilar todos los datos
                        datos_informe = {
                            'fecha': fecha,
                            'contrato': contrato,
                            'contratista': contratista,
                            'obra': f"{codigo_obra} {nombre_obra}" if codigo_obra and nombre_obra else "",
                            'codigo_obra': f"{codigo_obra}02" if codigo_obra else "",
                            'nombre_obra': nombre_obra,
                            'aprobacion': aprobacion,
                            'valor_contrato': valor_contrato,
                            'valor_pagado': valor_pagado,
                            'facturas': facturas_data,
                            'total_facturas': total_facturas,
                            # Por defecto, un nuevo certificado es 'Activo'
                            'estado': 'Activo',
                            'comentario_estado': None
                        }
                        
                        # Generar el informe
                        excel_data = generar_informe_excel(datos_informe, numero_certificado)
                        
                        if excel_data:
                            # Crear directorio para la obra si no existe
                            obra_dir = os.path.join(CERTIFICADOS_DIR, nombre_obra.replace("/", "_").replace("\\", "_"))
                            os.makedirs(obra_dir, exist_ok=True)
                            
                            # Guardar archivo físico con el nombre que incluye el número de certificado
                            filename = f"certificado_{numero_certificado:04d}.xlsx"
                            file_path = os.path.join(obra_dir, filename)
                            
                            with open(file_path, "wb") as f:
                                f.write(excel_data.getvalue())
                            
                            # Guardar en base de datos
                            certificado_id = guardar_certificado_db(
                                numero_certificado, obra_id, fecha, contrato, contratista,
                                valor_contrato, valor_pagado, total_facturas, facturas_data, file_path
                            )
                            
                            # Ofrecer el archivo para descargar
                            st.download_button(
                                label="📥 Descargar Informe",
                                data=excel_data,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                            st.success(f"✅ Certificado #{numero_certificado} para la obra '{nombre_obra}' generado correctamente!")
                        else:
                            st.error("❌ Error al generar el informe. Por favor intente nuevamente.")

elif menu_opcion == "📋 Ver Certificados":
    st.title("📋 Ver Certificados Generados")

    # --- NUEVO: PANEL DE BÚSQUEDA AVANZADA ---
    with st.expander("🔍 Búsqueda Avanzada"):
        st.markdown("Usa los siguientes filtros para refinar tu búsqueda.")
        
        # Obtener todas las obras para el multiselect
        obras_db = get_all_obras()
        opciones_obras = {f"{obra[1]} ({obra[2]})": obra[0] for obra in obras_db} # nombre (codigo): id

        col1, col2 = st.columns(2)
        with col1:
            filtro_obras_seleccionadas = st.multiselect(
                "Filtrar por obra(s):",
                options=list(opciones_obras.keys()),
                default=st.session_state.filtro_obras
            )
            filtro_estado_seleccionado = st.multiselect(
                "Filtrar por estado:",
                options=['Activo', 'Revertido', 'Cancelado'],
                default=st.session_state.filtro_estado
            )
        
        with col2:
            filtro_fecha_inicio = st.date_input(
                "Fecha de inicio:",
                value=st.session_state.filtro_fecha_inicio,
                format="YYYY-MM-DD"
            )
            filtro_fecha_fin = st.date_input(
                "Fecha de fin:",
                value=st.session_state.filtro_fecha_fin,
                format="YYYY-MM-DD"
            )
        
        filtro_contratista = st.text_input(
            "Buscar por Contratista:",
            value=st.session_state.filtro_contratista
        )

        col_boton_aplicar, col_boton_limpiar = st.columns(2)
        with col_boton_aplicar:
            if st.button("🔎 Aplicar Filtros", type="primary", use_container_width=True):
                st.session_state.filtros_aplicados = True
                st.session_state.filtro_obras = filtro_obras_seleccionadas
                st.session_state.filtro_estado = filtro_estado_seleccionado
                st.session_state.filtro_fecha_inicio = filtro_fecha_inicio
                st.session_state.filtro_fecha_fin = filtro_fecha_fin
                st.session_state.filtro_contratista = filtro_contratista
                st.rerun()
        
        with col_boton_limpiar:
            if st.button("🗑️ Limpiar Filtros", use_container_width=True):
                # Eliminar variables de estado para limpiar los filtros
                for key in ['filtros_aplicados', 'filtro_obras', 'filtro_estado', 'filtro_fecha_inicio', 'filtro_fecha_fin', 'filtro_contratista']:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()
    # --- FIN PANEL DE BÚSQUEDA AVANZADA ---

    st.markdown("---")
    
    # --- LÓGICA PARA OBTENER CERTIFICADOS ---
    if st.session_state.get('filtros_aplicados', False):
        # Preparar los IDs de las obras seleccionadas
        obras_ids_filtro = [opciones_obras[obra] for obra in st.session_state.filtro_obras]
        
        # Llamar a la nueva función de búsqueda con los filtros del session_state
        certificados = buscar_certificados_con_filtros(
            obras_ids=obras_ids_filtro if obras_ids_filtro else None,
            estados=st.session_state.filtro_estado if st.session_state.filtro_estado else None,
            fecha_inicio=st.session_state.filtro_fecha_inicio,
            fecha_fin=st.session_state.filtro_fecha_fin,
            contratista_texto=st.session_state.filtro_contratista if st.session_state.filtro_contratista else None
        )
    else:
        # Si no hay filtros aplicados, obtener todos los certificados
        certificados = get_certificados_by_obra()

    # --- EL RESTO DEL CÓDIGO DE LA PÁGINA PERMANECE IGUAL ---
    # (Mostrar la tabla, acciones rápidas, etc.)
    
    if certificados:
        st.write(f"### 📊 Certificados encontrados: {len(certificados)}")
        
        # ... (El código para mostrar el DataFrame se mantiene exactamente igual) ...
        # Crear una lista para almacenar los datos con acciones
        certificados_con_acciones = []
        
        # Procesar cada certificado para agregar acciones
        for cert in certificados:
            cert_dict = {
                'ID': cert[0], 'N° Certificado': cert[1], 'Obra Nombre': cert[13], 'Obra Codigo': cert[14],
                'Fecha': cert[3], 'Contratista': cert[5], 'Valor Contrato': cert[6], 'Valor Pagado': cert[7],
                'Total Facturas': cert[8], 'Fecha Generación': cert[10], 'Estado': cert[11]
            }
            certificados_con_acciones.append(cert_dict)
        
        df_mostrar = pd.DataFrame(certificados_con_acciones)
        
        df_mostrar['Valor Contrato'] = df_mostrar['Valor Contrato'].apply(lambda x: f"{x:,.2f}" if pd.notnull(x) and x != 0 else "0.00")
        df_mostrar['Valor Pagado'] = df_mostrar['Valor Pagado'].apply(lambda x: f"{x:,.2f}" if pd.notnull(x) and x != 0 else "0.00")
        df_mostrar['Total Facturas'] = df_mostrar['Total Facturas'].apply(lambda x: f"{x:,.2f}" if pd.notnull(x) and x != 0 else "0.00")
        
        def format_estado_with_emoji(val):
            if val in ['Revertido', 'Cancelado']: return f"🔴 {val}"
            elif val == 'Activo': return f"🟢 {val}"
            else: return f"⚪ {val}"
        
        df_mostrar['Estado'] = df_mostrar['Estado'].apply(format_estado_with_emoji)
        
        def highlight_estado(val):
            estado = val.split(' ', 1)[1] if ' ' in val else val
            color = 'red' if estado in ['Revertido', 'Cancelado'] else 'green' if estado == 'Activo' else ''
            return f'background-color: {color}'
        
        st.dataframe(df_mostrar.style.applymap(highlight_estado, subset=['Estado']), use_container_width=True, height=500)
        
        # ... (El resto del código de "Acciones Rápidas" y "Descargar Certificado" se mantiene igual) ...
        # Sección para seleccionar certificado directamente desde la tabla
        st.markdown("---")
        st.subheader("⚡ Acciones Rápidas")
        st.info("Para seleccionar un certificado, primero identifíquelo en la tabla superior. "
                "Después, ingrese su ID en el campo a continuación para realizar acciones sobre él.")
        
        col1, col2 = st.columns([3, 1])
        with col1:
            cert_id_input = st.text_input("Ingrese el ID del certificado (visible en la tabla):")
        with col2:
            st.write("")
            st.write("")
            if st.button("🔍 Buscar"):
                if cert_id_input.isdigit():
                    st.session_state.selected_cert_id = int(cert_id_input)
                    st.rerun()
                else:
                    st.warning("Por favor, ingrese un ID válido (número entero)")
        
        selected_cert = None
        if 'selected_cert_id' in st.session_state:
            cert_id = st.session_state.selected_cert_id
            for cert in certificados:
                if cert[0] == cert_id:
                    selected_cert = cert
                    break
            
            if selected_cert:
                st.success(f"Certificado seleccionado: #{selected_cert[1]} - {selected_cert[13]} ({selected_cert[14]}) - {selected_cert[3]}")
            else:
                st.error("Certificado no encontrado. Por favor, verifique el ID.")
                del st.session_state.selected_cert_id
        
        if selected_cert:
            cert_id = selected_cert[0]
            cert_numero = selected_cert[1]
            cert_obra = selected_cert[13]
            
            st.markdown("---")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("✏️ Editar Certificado", type="primary", use_container_width=True):
                    st.session_state.edit_cert_id = cert_id
                    go_to_page("editar")
            
            with col2:
                if st.button("🗑️ Eliminar Certificado", type="secondary", use_container_width=True):
                    st.session_state.delete_cert_id = cert_id
                    st.session_state.delete_cert_info = (cert_numero, cert_obra)
                    st.rerun()
            
            if 'delete_cert_id' in st.session_state and st.session_state.delete_cert_id == cert_id:
                st.warning(f"⚠️ ¿Está seguro que desea eliminar el certificado #{st.session_state.delete_cert_info[0]} de la obra {st.session_state.delete_cert_info[1]}?")
                st.warning("Esta acción no se puede deshacer y eliminará todas las facturas asociadas.")
                
                col3, col4 = st.columns(2)
                with col3:
                    if st.button("✅ Confirmar Eliminación", type="primary"):
                        delete_certificado(st.session_state.delete_cert_id)
                        st.success(f"Certificado #{st.session_state.delete_cert_info[0]} eliminado correctamente!")
                        del st.session_state.delete_cert_id
                        del st.session_state.delete_cert_info
                        if 'selected_cert_id' in st.session_state:
                            del st.session_state.selected_cert_id
                        st.rerun()
                
                with col4:
                    if st.button("❌ Cancelar", type="secondary"):
                        del st.session_state.delete_cert_id
                        del st.session_state.delete_cert_info
                        st.rerun()
        else:
            st.info("Ingrese el ID de un certificado para habilitar las acciones rápidas.")
        
        st.markdown("---")
        st.subheader("📥 Descargar Certificado")
        certificado_id_seleccion = st.selectbox(
            "Seleccione un certificado para descargar:",
            options=certificados,
            format_func=lambda x: f"#{x[1]} - {x[13]} ({x[14]}) - {x[3]} - [{x[11] if len(x) > 11 else 'N/A'}]"
        )
        
        if certificado_id_seleccion:
            certificado_id = certificado_id_seleccion[0]
            archivo_path = certificado_id_seleccion[9]
            
            if os.path.exists(archivo_path):
                with open(archivo_path, "rb") as file:
                    st.download_button(
                        label="📥 Descargar Certificado Seleccionado",
                        data=file,
                        file_name=os.path.basename(archivo_path),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            else:
                st.warning("Archivo no encontrado. Puede que haya sido movido o eliminado.")
    else:
        st.info("📭 No se encontraron certificados con los criterios seleccionados.")

        
elif menu_opcion == "✏️ Editar Certificado":
    st.title("✏️ Editar Certificado")
    
    # --- LÓGICA UNIFICADA PARA OBTENER EL CERTIFICADO A EDITAR ---
    certificado_id = None

    # Ruta A: Viniendo desde "Ver Certificados" a través del estado de la sesión
    if 'edit_cert_id' in st.session_state:
        certificado_id = st.session_state.edit_cert_id
        # Limpiamos el estado inmediatamente después de usarlo para evitar conflictos
        del st.session_state.edit_cert_id

    # Ruta B: Viniendo directamente desde el menú lateral
    else:
        st.subheader("Seleccionar un certificado para editar")
        todos_los_certificados = get_certificados_by_obra()
        
        if not todos_los_certificados:
            st.info("📭 No hay certificados disponibles para editar.")
            st.stop()
        
        certificado_seleccionado_tuple = st.selectbox(
            "Seleccione un certificado:",
            options=todos_los_certificados,
            format_func=lambda x: f"#{x[1]} - {x[13]} ({x[14]}) - {x[3]} - [{x[11] if len(x) > 11 else 'N/A'}]"
        )
        certificado_id = certificado_seleccionado_tuple[0]

    # --- OBTENER LOS DATOS Y PREPARAR VARIABLES ---
    # Ahora, sin importar la ruta, tenemos un certificado_id. Obtenemos los datos UNA SOLA VEZ.
    if certificado_id:
        certificado_data = get_certificado_by_id(certificado_id)
        if not certificado_data:
            st.error("No se pudo encontrar el certificado con el ID especificado.")
            st.stop()

        # Desempaquetamos los datos en variables con nombres claros para el resto de la sección
        # Índices: 0:id, 1:numero, 3:fecha, 4:contrato, 5:contratista, 6:valor_contrato, 7:valor_pagado, 8:total_facturas, 11:estado, 12:comentario, 13:obra_nombre, 14:obra_codigo
        numero_certificado = certificado_data[1]
        obra_nombre = certificado_data[13]
        obra_codigo = certificado_data[14]
        estado_actual = certificado_data[11] if len(certificado_data) > 11 else 'Activo'
        comentario_actual = certificado_data[12] if len(certificado_data) > 12 else ''
        
        # Obtener las facturas asociadas
        facturas_data = get_facturas_by_certificado_id(certificado_id)

        # --- AHORA CONSTRUIMOS LA INTERFAZ ---
        st.markdown(f"### 📝 Editando Certificado #{numero_certificado} - Obra: {obra_nombre} ({obra_codigo})")
        
        # Mostrar estado actual
        estado_color = "🔴" if estado_actual in ['Revertido', 'Cancelado'] else "🟢" if estado_actual == 'Activo' else "⚪"
        st.info(f"{estado_color} Estado actual: **{estado_actual}**")
        
        # Formulario de edición
        st.subheader("📄 Información del Certificado")
        col1, col2 = st.columns(2)
        with col1:
            fecha_edit = st.date_input("Fecha:", value=datetime.strptime(certificado_data[3], "%Y-%m-%d").date() if certificado_data[3] else datetime.now().date())
            contrato_edit = st.text_input("Contrato:", value=certificado_data[4] or "")
            valor_contrato_edit = st.number_input("Valor Contrato:", value=float(certificado_data[6] or 0.0), format="%.2f")
            # Campos para estado y comentario
            estado_edit = st.selectbox("Estado del Certificado:", 
                                      options=['Activo', 'Revertido', 'Cancelado'],
                                      index=['Activo', 'Revertido', 'Cancelado'].index(estado_actual) if estado_actual in ['Activo', 'Revertido', 'Cancelado'] else 0)
            comentario_estado_edit = st.text_area("Comentario sobre el estado (opcional):", 
                                                 value=comentario_actual or "",
                                                 height=100)
        
        with col2:
            contratista_edit = st.text_input("Contratista:", value=certificado_data[5] or "")
            valor_pagado_edit = st.number_input("Valor Pagado:", value=float(certificado_data[7] or 0.0), format="%.2f")
            total_facturas_edit = st.number_input("Total Facturas:", value=float(certificado_data[8] or 0.0), format="%.2f")
        
        st.markdown("---")
        st.subheader("📋 Facturas")
        
        # Editor de facturas
        facturas_edit_data = []
        for i, factura in enumerate(facturas_data):
            st.markdown(f"**Factura {i+1}**")
            col_prov, col_fact, col_imp, col_cod = st.columns(4)
            
            with col_prov:
                proveedor = st.text_input(f"Proveedor {i+1}", value=factura[0], key=f"edit_prov_{i}")
            with col_fact:
                numero_fact = st.text_input(f"Factura {i+1}", value=factura[1], key=f"edit_fact_{i}")
            with col_imp:
                importe = st.number_input(f"Importe {i+1}", value=float(factura[2]), format="%.2f", key=f"edit_imp_{i}")
            with col_cod:
                codigo = st.text_input(f"Código {i+1}", value=factura[3] or "", key=f"edit_cod_{i}")
            
            facturas_edit_data.append({
                'proveedor': proveedor,
                'factura': numero_fact,
                'importe': importe,
                'codigo': codigo
            })
        
        # Botón para guardar cambios
        st.markdown("---")
        if st.button("💾 Guardar Cambios", type="primary", use_container_width=True):
            # Validar datos
            if total_facturas_edit <= 0:
                st.error("El total de facturas debe ser mayor que 0")
            else:
                # Actualizar certificado (incluyendo estado y comentario)
                update_certificado(certificado_id, fecha_edit, contrato_edit, contratista_edit,
                                     valor_contrato_edit, valor_pagado_edit, total_facturas_edit,
                                     estado_edit, comentario_estado_edit)
                
                # Actualizar facturas
                update_facturas(certificado_id, facturas_edit_data)
                
                st.success("✅ Certificado actualizado correctamente!")
                
                # --- NUEVO: Navegamos de vuelta a la lista de certificados ---
                st.info("Redirigiendo a la lista de certificados...")
                go_to_page("ver")
    else:
        st.error("No se pudo determinar el certificado a editar.")